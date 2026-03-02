from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from app.core.database import get_db
from app.api.deps import get_current_admin
from app.models.order import Order, OrderItem, OrderStatus, STATUS_TRANSITIONS, STATUS_LABELS
from app.models.product import Product
from app.schemas.schemas import OrderOut, OrderStatusUpdate, OrderItemUpdate
from app.services.notifier import notify_user_status_change, notify_user_order_cancelled
import io, csv
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

router = APIRouter(prefix="/orders", tags=["orders"])

# ─── Цветовая палитра ────────────────────────────────────────────────────────
C_DARK       = "1A1A2E"   # Заголовок документа
C_ACCENT     = "E8A020"   # Золотой акцент (цвет бренда)
C_HEADER_BG  = "2D2D44"   # Шапка таблицы
C_HEADER_FG  = "FFFFFF"
C_ROW_ODD    = "F7F7FA"   # Светлые строки
C_ROW_EVEN   = "FFFFFF"
C_TOTAL_BG   = "EFF3FF"   # Строка итога
C_TOTAL_FG   = "1A1A2E"
C_META_LABEL = "6B6B80"   # Метка поля
C_META_VAL   = "1A1A2E"
C_BORDER     = "D0D0E0"

def _thin_border(sides="all"):
    t = Side(style="thin", color=C_BORDER)
    n = Side(style=None)
    if sides == "all":
        return Border(left=t, right=t, top=t, bottom=t)
    if sides == "bottom":
        return Border(bottom=t)
    if sides == "header":
        b = Side(style="medium", color=C_ACCENT)
        return Border(left=t, right=t, top=t, bottom=b)
    return Border()

def _header_font(size=10):
    return Font(name="Arial", bold=True, color=C_HEADER_FG, size=size)

def _cell_font(bold=False, color=C_META_VAL, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")


def _setup_print(ws, orientation="portrait"):
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_title_rows = None


def _write_doc_header(ws, title: str, subtitle: str, col_count: int):
    """Рисует красивую шапку документа в первых двух строках."""
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    cell.fill = _fill(C_DARK)
    cell.alignment = _center()

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.font = Font(name="Arial", color=C_ACCENT, size=10, bold=False)
    sub.fill = _fill(C_DARK)
    sub.alignment = _center()


def _write_meta_block(ws, meta: dict, start_row: int, col_count: int) -> int:
    """Пишет блок метаданных (ключ: значение). Возвращает следующую строку."""
    ws.row_dimensions[start_row].height = 6  # небольшой отступ

    row = start_row + 1
    items = list(meta.items())
    # Два столбца: пары key/value рядом
    pairs_per_row = 2
    for i in range(0, len(items), pairs_per_row):
        ws.row_dimensions[row].height = 18
        for j, (k, v) in enumerate(items[i:i+pairs_per_row]):
            col_k = 1 + j * 3
            col_v = col_k + 1
            lbl = ws.cell(row=row, column=col_k, value=k)
            lbl.font = Font(name="Arial", color=C_META_LABEL, size=9)
            lbl.alignment = _right()

            val = ws.cell(row=row, column=col_v, value=v)
            val.font = Font(name="Arial", color=C_META_VAL, size=9, bold=True)
            val.alignment = _left()
        row += 1

    ws.row_dimensions[row].height = 6  # отступ снизу
    return row + 1


def _write_table(ws, headers: list, rows: list, start_row: int,
                 col_widths: list = None, money_cols: set = None,
                 num_cols: set = None) -> int:
    """Рисует таблицу с заголовком и данными. Возвращает следующую строку."""
    money_cols = money_cols or set()
    num_cols = num_cols or set()

    # Шапка таблицы
    ws.row_dimensions[start_row].height = 22
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = _header_font()
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border = _thin_border("header")

    # Строки данных
    for ri, row_data in enumerate(rows):
        r = start_row + 1 + ri
        ws.row_dimensions[r].height = 18
        bg = C_ROW_ODD if ri % 2 == 0 else C_ROW_EVEN
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = _fill(bg)
            c.border = _thin_border()
            if ci in money_cols:
                c.number_format = '#,##0.00 ₽'
                c.alignment = _right()
                c.font = Font(name="Arial", size=10)
            elif ci in num_cols:
                c.alignment = _center()
                c.font = Font(name="Arial", size=10)
            else:
                c.alignment = _left()
                c.font = Font(name="Arial", size=10)

    return start_row + 1 + len(rows)


def _write_total_row(ws, row: int, col_count: int, label_col: int,
                     total_col: int, total_val):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_col)
    lbl = ws.cell(row=row, column=1, value="ИТОГО")
    lbl.font = Font(name="Arial", bold=True, color=C_TOTAL_FG, size=10)
    lbl.fill = _fill(C_TOTAL_BG)
    lbl.alignment = _right()
    lbl.border = _thin_border()

    for ci in range(label_col + 1, col_count + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = _fill(C_TOTAL_BG)
        c.border = _thin_border()

    tot = ws.cell(row=row, column=total_col, value=total_val)
    tot.font = Font(name="Arial", bold=True, color=C_TOTAL_FG, size=10)
    tot.fill = _fill(C_TOTAL_BG)
    tot.number_format = '#,##0.00 ₽'
    tot.alignment = _right()
    tot.border = _thin_border()


def _autofit_columns(ws, col_widths: list):
    for ci, (min_w, max_w) in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(min_w, min(max_w, min_w))


# ─── Router helpers ──────────────────────────────────────────────────────────

def _order_query():
    return (
        select(Order)
        .options(
            selectinload(Order.user),
            selectinload(Order.items).selectinload(OrderItem.product),
        )
        .order_by(Order.created_at.desc())
    )


# ─── Routes ──────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[OrderOut])
async def get_orders(
    status: OrderStatus | None = None,
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    q = _order_query()
    if status:
        q = q.where(Order.status == status)
    result = await db.execute(q)
    return result.scalars().all()


@router.get("/export/pending")
async def export_pending_orders(
    format: str = "xlsx",
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    result = await db.execute(_order_query().where(Order.status == OrderStatus.PENDING))
    orders = result.scalars().all()

    # Сводка по товарам
    summary: dict[str, dict] = {}
    for o in orders:
        for item in o.items:
            name = item.product.name if item.product else f"Товар #{item.product_id}"
            if name not in summary:
                summary[name] = {"qty": 0, "price": item.price_at_order}
            summary[name]["qty"] += item.quantity

    summary_rows = sorted(summary.items())

    if format == "csv":
        output = io.StringIO()
        output.write("=== СВОДНЫЙ СПИСОК ТОВАРОВ ===\n")
        output.write("Товар,Количество,Цена за шт.,Сумма\n")
        for name, v in summary_rows:
            output.write(f"{name},{v['qty']},{v['price']},{v['price'] * v['qty']}\n")
        output.write("\n=== ДЕТАЛИ ЗАКАЗОВ ===\n")
        output.write("ID,Клиент,Telegram,Телефон,Товары,Комментарий,Дата\n")
        for o in orders:
            items_str = " | ".join(
                f"{i.product.name if i.product else i.product_id} x{i.quantity}"
                for i in o.items
            )
            phone = o.user.phone or ""
            tg = f"@{o.user.username}" if o.user.username else str(o.user.telegram_id)
            name = o.user.full_name or o.user.username or str(o.user.telegram_id)
            output.write(f'{o.id},{name},{tg},{phone},"{items_str}","{o.comment or ""}",{o.created_at.strftime("%Y-%m-%d %H:%M")}\n')
        output.seek(0)
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=pending_orders.csv"})

    # ── XLSX ──────────────────────────────────────────────────────────────────
    from datetime import datetime
    wb = openpyxl.Workbook()

    # ── Лист 1: Сводка товаров ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Сводка товаров"
    _setup_print(ws1)

    total_qty = sum(v["qty"] for _, v in summary_rows)
    total_sum = sum(v["qty"] * v["price"] for _, v in summary_rows)
    subtitle = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}  ·  Заказов: {len(orders)}  ·  Позиций: {len(summary_rows)}"
    _write_doc_header(ws1, "СВОДНЫЙ СПИСОК ТОВАРОВ — На подтверждении", subtitle, 4)

    headers1 = ["Товар", "Кол-во", "Цена за шт.", "Сумма"]
    data1 = [
        (name, v["qty"], v["price"], v["price"] * v["qty"])
        for name, v in summary_rows
    ]
    next_row = _write_table(ws1, headers1, data1, start_row=4,
                            money_cols={3, 4}, num_cols={2})
    _write_total_row(ws1, next_row, 4, label_col=2, total_col=4, total_val=total_sum)

    # Ширины столбцов для А4
    col_widths1 = [(36, 36), (10, 10), (14, 14), (16, 16)]
    for ci, (w, _) in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Лист 2: Детали заказов ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Заказы")
    _setup_print(ws2, orientation="landscape")

    _write_doc_header(ws2, "ДЕТАЛИ ЗАКАЗОВ — На подтверждении", subtitle, 7)

    headers2 = ["№ заказа", "Клиент", "Telegram", "Телефон", "Товары", "Комментарий", "Дата"]
    data2 = []
    for o in orders:
        items_str = "\n".join(
            f"• {i.product.name if i.product else i.product_id} × {i.quantity}"
            for i in o.items
        )
        tg = f"@{o.user.username}" if o.user.username else str(o.user.telegram_id)
        client_name = o.user.full_name or o.user.username or str(o.user.telegram_id)
        phone = o.user.phone or "—"
        data2.append((
            f"#{o.id}",
            client_name,
            tg,
            phone,
            items_str,
            o.comment or "",
            o.created_at.strftime("%d.%m.%Y %H:%M"),
        ))

    next_row2 = _write_table(ws2, headers2, data2, start_row=4, num_cols={1})

    # Высота строк под перенос текста
    for ri in range(len(data2)):
        r = 4 + 1 + ri
        lines = max(
            data2[ri][4].count("\n") + 1,
            1
        )
        ws2.row_dimensions[r].height = max(18, 16 * lines)

    # Ширины для ландшафтного А4
    col_widths2 = [9, 20, 16, 14, 38, 20, 14]
    for ci, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pending_orders.xlsx"})


@router.get("/{order_id}", response_model=OrderOut)
async def get_order(order_id: int, db: AsyncSession = Depends(get_db), admin=Depends(get_current_admin)):
    result = await db.execute(_order_query().where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.patch("/{order_id}/status", response_model=OrderOut)
async def update_order_status(
    order_id: int,
    data: OrderStatusUpdate,
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    result = await db.execute(_order_query().where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    allowed = STATUS_TRANSITIONS.get(order.status, [])
    if data.status not in allowed:
        raise HTTPException(status_code=400,
            detail=f"Нельзя перейти из '{order.status}' в '{data.status}'")

    order.status = data.status
    await db.commit()

    if data.status == OrderStatus.CANCELLED:
        await notify_user_order_cancelled(order.user.telegram_id, order.id)
    else:
        await notify_user_status_change(order.user.telegram_id, order.id, data.status)

    result2 = await db.execute(_order_query().where(Order.id == order_id))
    return result2.scalar_one()


@router.patch("/{order_id}/items/{item_id}", response_model=OrderOut)
async def update_order_item(
    order_id: int,
    item_id: int,
    data: OrderItemUpdate,
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    result = await db.execute(select(OrderItem).where(OrderItem.id == item_id, OrderItem.order_id == order_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    if data.quantity <= 0:
        await db.delete(item)
    else:
        item.quantity = data.quantity
    await db.commit()
    result2 = await db.execute(_order_query().where(Order.id == order_id))
    return result2.scalar_one()


@router.delete("/{order_id}/items/{item_id}", response_model=OrderOut)
async def delete_order_item(
    order_id: int,
    item_id: int,
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    result = await db.execute(select(OrderItem).where(OrderItem.id == item_id, OrderItem.order_id == order_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    await db.delete(item)
    await db.commit()
    result2 = await db.execute(_order_query().where(Order.id == order_id))
    return result2.scalar_one()


@router.get("/{order_id}/export")
async def export_single_order(
    order_id: int,
    format: str = "xlsx",
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    result = await db.execute(_order_query().where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    total = sum(i.price_at_order * i.quantity for i in order.items)
    client_name = order.user.full_name or order.user.username or str(order.user.telegram_id)
    tg = f"@{order.user.username}" if order.user.username else str(order.user.telegram_id)

    if format == "csv":
        output = io.StringIO()
        output.write(f"Заказ №,{order.id}\n")
        output.write(f"Клиент,{client_name}\n")
        output.write(f"Telegram,{tg}\n")
        output.write(f"Телефон,{order.user.phone or ''}\n")
        output.write(f"Статус,{STATUS_LABELS[order.status]}\n")
        output.write(f"Комментарий,{order.comment or ''}\n")
        output.write(f"Дата,{order.created_at.strftime('%Y-%m-%d %H:%M')}\n\n")
        output.write("Товар,Количество,Цена за шт.,Сумма\n")
        for item in order.items:
            name = item.product.name if item.product else f"#{item.product_id}"
            output.write(f"{name},{item.quantity},{item.price_at_order},{item.price_at_order * item.quantity}\n")
        output.write(f",,ИТОГО,{total}\n")
        output.seek(0)
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=order_{order_id}.csv"})

    # ── XLSX ──────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заказ #{order_id}"
    _setup_print(ws)

    COL_COUNT = 4

    # Заголовок
    _write_doc_header(
        ws,
        f"ЗАКАЗ № {order_id}",
        f"{STATUS_LABELS[order.status]}  ·  {order.created_at.strftime('%d.%m.%Y %H:%M')}",
        COL_COUNT,
    )

    # Блок метаданных
    phone_val = order.user.phone or "—"
    meta = {
        "Клиент:": client_name,
        "Telegram:": tg,
        "Телефон:": phone_val,
        "Статус:": STATUS_LABELS[order.status],
        "Дата:": order.created_at.strftime("%d.%m.%Y %H:%M"),
        "Комментарий:": order.comment or "—",
    }
    table_start = _write_meta_block(ws, meta, start_row=3, col_count=COL_COUNT)

    # Таблица товаров
    headers = ["Наименование товара", "Кол-во", "Цена за шт.", "Сумма"]
    rows = [
        (
            item.product.name if item.product else f"#{item.product_id}",
            item.quantity,
            item.price_at_order,
            item.price_at_order * item.quantity,
        )
        for item in order.items
    ]
    next_row = _write_table(ws, headers, rows, start_row=table_start,
                            money_cols={3, 4}, num_cols={2})
    _write_total_row(ws, next_row, COL_COUNT, label_col=2, total_col=4, total_val=total)

    # Ширины столбцов для А4 (портрет)
    col_widths = [38, 10, 16, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=order_{order_id}.xlsx"},
    )
