from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from app.core.database import get_db
from app.api.deps import get_current_admin
from app.models.product import Product
from app.models.category import Category
from app.schemas.schemas import ProductCreate, ProductUpdate, ProductOut
import io, csv
import openpyxl

router = APIRouter(prefix="/products", tags=["products"])


def _product_query():
    return select(Product).options(
        selectinload(Product.category).selectinload(Category.children),
        selectinload(Product.category).selectinload(Category.parent),
    )


@router.get("/", response_model=list[ProductOut])
async def get_products(
    category_id: int | None = None,
    visible_only: bool = False,
    db: AsyncSession = Depends(get_db),
):
    q = _product_query()
    if category_id is not None:
        q = q.where(Product.category_id == category_id)
    if visible_only:
        q = q.where(Product.is_visible == True)
    result = await db.execute(q)
    return result.scalars().all()


@router.get("/{product_id}", response_model=ProductOut)
async def get_product(product_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(_product_query().where(Product.id == product_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    return p


@router.post("/", response_model=ProductOut)
async def create_product(
    data: ProductCreate,
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    p = Product(**data.model_dump())
    db.add(p)
    await db.commit()
    result = await db.execute(_product_query().where(Product.id == p.id))
    return result.scalar_one()


@router.put("/{product_id}", response_model=ProductOut)
async def update_product(
    product_id: int,
    data: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    for k, v in data.model_dump().items():
        setattr(p, k, v)
    await db.commit()
    result = await db.execute(_product_query().where(Product.id == product_id))
    return result.scalar_one()


@router.delete("/{product_id}")
async def delete_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    from app.models.order import OrderItem
    from sqlalchemy import func as sqlfunc

    result = await db.execute(select(Product).where(Product.id == product_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")

    # Считаем сколько позиций заказов ссылаются на товар
    usage = await db.execute(
        select(sqlfunc.count(OrderItem.id)).where(OrderItem.product_id == product_id)
    )
    orders_count = usage.scalar() or 0

    await db.delete(p)
    await db.commit()

    return {
        "ok": True,
        "orders_affected": orders_count,
        "message": (
            f"Товар удалён. В {orders_count} позициях заказов он будет отображаться как «Удалённый товар»."
            if orders_count > 0 else "Товар удалён."
        )
    }


# ─── Вспомогательные функции импорта ─────────────────────────────────────────

def _parse_price(val) -> float | None:
    """Конвертирует значение в float, возвращает None если не получилось."""
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").replace(" ", "").strip())
    except (ValueError, TypeError):
        return None


def _parse_visible(val) -> bool:
    """1 / '1' / 'да' / 'yes' / 'true' → True, всё остальное → False."""
    if val is None:
        return False
    return str(val).strip().lower() in ("1", "да", "yes", "true", "+")


async def _get_or_create_category(name: str, db: AsyncSession) -> int | None:
    """Ищет категорию по имени, создаёт если не найдена. Возвращает id."""
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    result = await db.execute(select(Category).where(Category.name == name))
    cat = result.scalar_one_or_none()
    if not cat:
        cat = Category(name=name, is_visible=True)
        db.add(cat)
        await db.flush()  # получаем id без commit
    return cat.id


def _detect_format(headers: list[str]) -> str:
    """
    Определяет формат файла по заголовкам.
    Возвращает 'extended' (формат с Разделами) или 'simple' (старый формат).
    """
    h = [str(h).strip().lower() if h else "" for h in headers]
    if "раздел 1" in h or "наименование" in h:
        return "extended"
    return "simple"


def _normalize_row_extended(row: dict) -> dict | None:
    """
    Парсит строку в формате с колонками:
    ID, Активно, Раздел 1, Раздел 2, Артикул, Наименование,
    Описание, Ед. изм, Вес, Объем, Диаметр, Размер, Цвет,
    Цена, Старая цена, Изображение, Ссылка
    """
    name = row.get("Наименование") or row.get("наименование")
    price = _parse_price(row.get("Цена") or row.get("цена"))
    if not name or price is None:
        return None

    # Категория: приоритет у Раздел 2, затем Раздел 1
    category_name = (
        row.get("Раздел 2") or row.get("раздел 2") or
        row.get("Раздел 1") or row.get("раздел 1")
    )

    # Видимость: колонка «Активно»
    active_val = row.get("Активно") or row.get("активно")
    is_visible = _parse_visible(active_val) if active_val is not None else True

    photo_url = row.get("Изображение") or row.get("изображение") or None
    if photo_url:
        photo_url = str(photo_url).strip() or None

    description = row.get("Описание") or row.get("описание") or None
    if description:
        description = str(description).strip() or None

    unit = row.get("Ед. изм") or row.get("ед. изм") or row.get("ед.изм") or None
    if unit:
        unit = str(unit).strip() or None

    weight = row.get("Вес") or row.get("вес") or row.get("Объем") or row.get("объем") or None
    if weight:
        weight = str(weight).strip() or None

    return {
        "name": str(name).strip(),
        "description": description,
        "price": price,
        "unit": unit,
        "weight": weight,
        "photo_url": photo_url,
        "is_visible": is_visible,
        "category_name": str(category_name).strip() if category_name else None,
    }


def _normalize_row_simple(row: dict) -> dict | None:
    """Старый простой формат: name, price, description, photo_url."""
    name = row.get("name") or row.get("название") or row.get("наименование")
    price = _parse_price(row.get("price") or row.get("цена"))
    if not name or price is None:
        return None

    photo_url = row.get("photo_url") or row.get("фото") or row.get("изображение") or None
    if photo_url:
        photo_url = str(photo_url).strip() or None

    description = row.get("description") or row.get("описание") or None
    if description:
        description = str(description).strip() or None

    unit = row.get("unit") or row.get("Ед. изм") or row.get("ед. изм") or None
    if unit:
        unit = str(unit).strip() or None

    weight = row.get("weight") or row.get("Вес") or row.get("вес") or None
    if weight:
        weight = str(weight).strip() or None

    return {
        "name": str(name).strip(),
        "description": description,
        "price": price,
        "unit": unit,
        "weight": weight,
        "photo_url": photo_url,
        "is_visible": True,
        "category_name": None,
    }


@router.post("/import")
async def import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin=Depends(get_current_admin),
):
    content = await file.read()
    raw_rows: list[dict] = []

    if file.filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
        raw_rows = list(reader)
        fmt = _detect_format(list(raw_rows[0].keys()) if raw_rows else [])

    elif file.filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        # Берём первый непустой лист
        ws = None
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row > 1:
                ws = sheet
                break
        if ws is None:
            raise HTTPException(status_code=400, detail="Файл не содержит данных")

        headers = [cell.value for cell in ws[1]]
        fmt = _detect_format(headers)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            raw_rows.append(dict(zip(headers, row)))
    else:
        raise HTTPException(status_code=400, detail="Поддерживаются только CSV и XLSX")

    # Нормализуем строки
    parsed = []
    for row in raw_rows:
        if fmt == "extended":
            result = _normalize_row_extended(row)
        else:
            result = _normalize_row_simple(row)
        if result:
            parsed.append(result)

    if not parsed:
        raise HTTPException(status_code=400, detail="Не найдено ни одной валидной строки. Проверьте колонки: «Наименование» и «Цена»")

    # Кэш категорий чтобы не делать лишние запросы
    category_cache: dict[str, int] = {}
    created = 0
    skipped = 0

    for item in parsed:
        cat_name = item.pop("category_name")
        category_id = None

        if cat_name:
            if cat_name not in category_cache:
                category_cache[cat_name] = await _get_or_create_category(cat_name, db)
            category_id = category_cache[cat_name]

        try:
            p = Product(
                name=item["name"],
                description=item.get("description"),
                price=item["price"],
                unit=item.get("unit"),
                weight=item.get("weight"),
                photo_url=item.get("photo_url"),
                is_visible=item.get("is_visible", True),
                category_id=category_id,
            )
            db.add(p)
            created += 1
        except Exception:
            skipped += 1

    await db.commit()

    return {
        "imported": created,
        "skipped": skipped,
        "categories_created": len(category_cache),
        "format_detected": fmt,
    }
